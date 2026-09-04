import argparse
import os
import re
import time
from pathlib import Path
from typing import Dict, List, Tuple
import numpy as np
import torch
import torch.distributed as dist
import torch.nn as nn
import torch.nn.functional as F
from torch.cuda.amp import GradScaler, autocast
from torch.nn.parallel import DistributedDataParallel as DDP
from torch.utils.data import DataLoader, DistributedSampler
from data import PairedBratsValDataset, UnpairedBratsDataset, build_transform
from models import NLayerDiscriminator, ResnetGenerator, get_norm_layer, init_weights
from afd_utils import AdaptiveFrequencyDecomposition, adaptive_frequency_consistency_loss
from utils_ import GANLoss, ImagePool, append_csv, get_linear_decay_lr, is_main_process, psnr_per_image, seed_everything, seed_worker, set_optimizer_lr, set_requires_grad, ssim_per_image, tensor_to_01, unwrap_model

def parse_args():
    parser = argparse.ArgumentParser('CycleGAN with FDDT')
    parser.add_argument('--dataroot', type=str, default='')
    parser.add_argument('--checkpoints_dir', type=str, default='')
    parser.add_argument('--name', type=str, default='fddt_cyclegan')
    parser.add_argument('--source_domain', type=str, default='')
    parser.add_argument('--target_domain', type=str, default='')
    parser.add_argument('--eval_phase', type=str, default='val')
    parser.add_argument('--eval_pairing', type=str, default='key', choices=['key', 'order'])
    parser.add_argument('--case_id_regex', type=str, default='')
    parser.add_argument('--input_nc', type=int, default=1)
    parser.add_argument('--output_nc', type=int, default=1)
    parser.add_argument('--load_size', type=int, default=0)
    parser.add_argument('--crop_size', type=int, default=0)
    parser.add_argument('--no_flip', action='store_true', default=True)
    parser.add_argument('--num_workers', type=int, default=8)
    parser.add_argument('--prefetch_factor', type=int, default=4)
    parser.add_argument('--ngf', type=int, default=64)
    parser.add_argument('--ndf', type=int, default=64)
    parser.add_argument('--netG_blocks', type=int, default=9)
    parser.add_argument('--n_layers_D', type=int, default=3)
    parser.add_argument('--norm', type=str, default='instance', choices=['instance', 'batch', 'none'])
    parser.add_argument('--gan_mode', type=str, default='lsgan', choices=['lsgan', 'vanilla'])
    parser.add_argument('--pool_size', type=int, default=50)
    parser.add_argument('--lambda_A', type=float, default=10.0)
    parser.add_argument('--lambda_B', type=float, default=10.0)
    parser.add_argument('--lambda_identity', type=float, default=0.1)
    parser.add_argument('--batch_size', type=int, default=1)
    parser.add_argument('--val_batch_size', type=int, default=0)
    parser.add_argument('--lr', type=float, default=0.0001)
    parser.add_argument('--beta1', type=float, default=0.5)
    parser.add_argument('--beta2', type=float, default=0.999)
    parser.add_argument('--n_epochs', type=int, default=100)
    parser.add_argument('--n_epochs_decay', type=int, default=100)
    parser.add_argument('--grad_clip', type=float, default=0.0)
    parser.add_argument('--amp', action='store_true')
    parser.add_argument('--channels_last', action='store_true')
    parser.add_argument('--seed', type=int, default=42)
    parser.add_argument('--patience', type=int, default=20)
    parser.add_argument('--early_metric', type=str, default='both', choices=['ssim', 'psnr', 'both'])
    parser.add_argument('--min_delta_ssim', type=float, default=1e-05)
    parser.add_argument('--min_delta_psnr', type=float, default=0.001)
    parser.add_argument('--local_rank', type=int, default=-1)
    parser.add_argument('--local-rank', type=int, default=-1)
    parser.add_argument('--resume', type=str, default='')
    parser.add_argument('--lambda_freq', type=float, default=None)
    parser.add_argument('--afd_embed_dim', type=int, default=None)
    parser.add_argument('--afd_num_heads', type=int, default=None)
    parser.add_argument('--afd_high_min', type=float, default=None)
    parser.add_argument('--afd_high_max', type=float, default=None)
    parser.add_argument('--afd_low_min', type=float, default=None)
    parser.add_argument('--afd_low_max', type=float, default=None)
    parser.add_argument('--afd_input_size', type=int, default=None)
    parser.add_argument('--afd_resume', type=str, default='')
    parser.add_argument('--fddt_bidir', action='store_true')
    args = parser.parse_args()
    required = {'dataroot': args.dataroot, 'checkpoints_dir': args.checkpoints_dir, 'source_domain': args.source_domain, 'target_domain': args.target_domain, 'case_id_regex': args.case_id_regex, 'lambda_freq': args.lambda_freq, 'afd_embed_dim': args.afd_embed_dim, 'afd_num_heads': args.afd_num_heads, 'afd_high_min': args.afd_high_min, 'afd_high_max': args.afd_high_max, 'afd_low_min': args.afd_low_min, 'afd_low_max': args.afd_low_max, 'afd_input_size': args.afd_input_size}
    missing = [key for key, value in required.items() if value in (None, '')]
    if missing:
        parser.error('Required settings are empty: ' + ', '.join(missing))
    return args

def setup_ddp():
    if 'RANK' in os.environ and 'WORLD_SIZE' in os.environ:
        rank = int(os.environ['RANK'])
        world_size = int(os.environ['WORLD_SIZE'])
        local_rank = int(os.environ.get('LOCAL_RANK', 0))
        torch.cuda.set_device(local_rank)
        dist.init_process_group(backend='nccl', init_method='env://')
    else:
        rank, world_size, local_rank = (0, 1, 0)
    return (rank, world_size, local_rank)

def cleanup_ddp():
    if dist.is_available() and dist.is_initialized():
        dist.barrier()
        dist.destroy_process_group()

def ddp_all_gather_object(obj):
    if dist.is_available() and dist.is_initialized():
        gathered = [None for _ in range(dist.get_world_size())]
        dist.all_gather_object(gathered, obj)
        return gathered
    return [obj]

def case_id_from_any_path_or_key(x: str, case_id_regex: str='') -> str:
    stem = Path(str(x)).stem
    if case_id_regex:
        match = re.search(case_id_regex, stem)
        if match is None:
            raise ValueError(f'case_id_regex did not match evaluation file: {stem}')
        return match.group(1) if match.groups() else match.group(0)
    if '_' in stem:
        parts = stem.split('_')
        if len(parts) >= 2 and parts[0].startswith('BraTS'):
            return f'{parts[0]}_{parts[1]}'
        if len(parts) >= 2:
            return '_'.join(parts[:2])
        return stem
    if '-' in stem:
        parts = stem.split('-')
        if len(parts) >= 4 and parts[0] == 'BraTS':
            return '-'.join(parts[:4])
        if len(parts) >= 3:
            return '-'.join(parts[:-2])
        return stem
    return stem

def slice_id_from_any_path_or_key(x: str) -> str:
    stem = Path(str(x)).stem
    return stem

def get_fddt_weight(args, epoch: int) -> float:
    return float(args.lambda_freq)

def compute_case_metrics_from_slice_entries(slice_entries: List[Dict]) -> Tuple[List[Dict], Dict]:
    unique = {}
    for item in slice_entries:
        sid = item['slice_id']
        if sid not in unique:
            unique[sid] = item
    grouped: Dict[str, List[Dict]] = {}
    for item in unique.values():
        grouped.setdefault(item['case_id'], []).append(item)
    case_rows = []
    for case_id in sorted(grouped.keys()):
        items = grouped[case_id]
        ssim_arr = np.asarray([x['ssim'] for x in items], dtype=np.float64)
        psnr_arr = np.asarray([x['psnr'] for x in items], dtype=np.float64)
        row = {'case_id': case_id, 'n_slices': int(len(items)), 'ssim_mean': float(ssim_arr.mean()), 'ssim_std_slice': float(ssim_arr.std(ddof=1)) if len(ssim_arr) > 1 else 0.0, 'psnr_mean': float(psnr_arr.mean()), 'psnr_std_slice': float(psnr_arr.std(ddof=1)) if len(psnr_arr) > 1 else 0.0}
        case_rows.append(row)
    case_ssim = np.asarray([x['ssim_mean'] for x in case_rows], dtype=np.float64)
    case_psnr = np.asarray([x['psnr_mean'] for x in case_rows], dtype=np.float64)
    summary = {'n_cases': int(len(case_rows)), 'n_slices': int(len(unique)), 'case_ssim_mean': float(case_ssim.mean()) if len(case_ssim) else 0.0, 'case_ssim_std': float(case_ssim.std(ddof=1)) if len(case_ssim) > 1 else 0.0, 'case_psnr_mean': float(case_psnr.mean()) if len(case_psnr) else 0.0, 'case_psnr_std': float(case_psnr.std(ddof=1)) if len(case_psnr) > 1 else 0.0}
    return (case_rows, summary)

def save_case_metrics_to_xlsx(xlsx_path: str, epoch: int, case_rows: List[Dict], summary: Dict):
    try:
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Font, Alignment
    except Exception as e:
        raise RuntimeError('openpyxl is required to save xlsx files: pip install openpyxl') from e
    xlsx_path = Path(xlsx_path)
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    if xlsx_path.exists():
        wb = load_workbook(str(xlsx_path))
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = 'case_level'
    if 'case_level' not in wb.sheetnames:
        wb.create_sheet('case_level')
    if 'epoch_summary' not in wb.sheetnames:
        wb.create_sheet('epoch_summary')
    ws_case = wb['case_level']
    ws_sum = wb['epoch_summary']
    case_header = ['epoch', 'case_id', 'n_slices', 'ssim_mean', 'ssim_std_slice', 'psnr_mean', 'psnr_std_slice']
    sum_header = ['epoch', 'n_cases', 'n_slices', 'case_ssim_mean', 'case_ssim_std', 'case_psnr_mean', 'case_psnr_std']
    if ws_case.max_row == 1 and ws_case.cell(1, 1).value is None:
        ws_case.append(case_header)
    if ws_sum.max_row == 1 and ws_sum.cell(1, 1).value is None:
        ws_sum.append(sum_header)

    def remove_epoch_rows(ws):
        rows_to_delete = []
        for r in range(2, ws.max_row + 1):
            if ws.cell(r, 1).value == epoch:
                rows_to_delete.append(r)
        for r in reversed(rows_to_delete):
            ws.delete_rows(r, 1)
    remove_epoch_rows(ws_case)
    remove_epoch_rows(ws_sum)
    for row in case_rows:
        ws_case.append([epoch, row['case_id'], row['n_slices'], row['ssim_mean'], row['ssim_std_slice'], row['psnr_mean'], row['psnr_std_slice']])
    ws_sum.append([epoch, summary['n_cases'], summary['n_slices'], summary['case_ssim_mean'], summary['case_ssim_std'], summary['case_psnr_mean'], summary['case_psnr_std']])
    for ws in [ws_case, ws_sum]:
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        ws.freeze_panes = 'A2'
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                val = cell.value
                if val is not None:
                    max_len = max(max_len, len(str(val)))
            ws.column_dimensions[col_letter].width = min(max_len + 2, 28)
    wb.save(str(xlsx_path))

@torch.inference_mode()
def validate_full_case_level(netG_A, val_loader, device, args):
    netG_A.eval()
    local_entries = []
    for batch in val_loader:
        real_A = batch['A'].to(device, non_blocking=True)
        real_B = batch['B'].to(device, non_blocking=True)
        if args.channels_last:
            real_A = real_A.contiguous(memory_format=torch.channels_last)
            real_B = real_B.contiguous(memory_format=torch.channels_last)
        with autocast(enabled=args.amp):
            fake_B = netG_A(real_A)
        fake01 = tensor_to_01(fake_B)
        realB01 = tensor_to_01(real_B)
        ssim_vals = ssim_per_image(fake01, realB01).detach().float().cpu().numpy()
        psnr_vals = psnr_per_image(fake01, realB01).detach().float().cpu().numpy()
        if 'key' in batch:
            ids = batch['key']
        elif 'A_path' in batch:
            ids = batch['A_path']
        else:
            ids = [str(i) for i in range(len(ssim_vals))]
        for i in range(len(ssim_vals)):
            x = ids[i]
            local_entries.append({'case_id': case_id_from_any_path_or_key(x, args.case_id_regex), 'slice_id': slice_id_from_any_path_or_key(x), 'ssim': float(ssim_vals[i]), 'psnr': float(psnr_vals[i])})
    gathered = ddp_all_gather_object(local_entries)
    all_entries = []
    for part in gathered:
        all_entries.extend(part)
    case_rows, summary = compute_case_metrics_from_slice_entries(all_entries)
    netG_A.train()
    return (case_rows, summary)

def save_G_A_only(path: str, netG_A) -> None:
    os.makedirs(os.path.dirname(path), exist_ok=True)
    torch.save(unwrap_model(netG_A).state_dict(), path)

def main():
    args = parse_args()
    rank, world_size, local_rank = setup_ddp()
    device = torch.device(f'cuda:{local_rank}' if torch.cuda.is_available() else 'cpu')
    seed_everything(args.seed + rank)
    torch.backends.cudnn.benchmark = True
    try:
        torch.set_float32_matmul_precision('high')
    except Exception:
        pass
    exp_dir = os.path.join(args.checkpoints_dir, args.name)
    xlsx_path = os.path.join(exp_dir, 'case_level_metrics.xlsx')
    if is_main_process():
        os.makedirs(exp_dir, exist_ok=True)
        print(f'[DDP] world_size={world_size}')
        print(f'[DDP] per_gpu_train_batch={args.batch_size}')
        print(f'[DDP] global_train_batch={args.batch_size * world_size}')
        print(f'[DDP] per_gpu_val_batch={(args.val_batch_size if args.val_batch_size > 0 else args.batch_size * 4)}')
        print(f'[Data] train source: {args.dataroot}/train/{args.source_domain}')
        print(f'[Data] eval source: {args.dataroot}/{args.eval_phase}/{args.source_domain}')
        print(f'[Input] original size, no resize, no crop, no augmentation')
        print(f'[Save] {exp_dir}')
        print(f'[Excel] {xlsx_path}')
    train_tf = build_transform(phase='train', input_nc=args.input_nc, load_size=args.load_size, crop_size=args.crop_size, no_flip=True)
    val_tf = build_transform(phase='val', input_nc=args.input_nc, load_size=args.load_size, crop_size=args.crop_size, no_flip=True)
    train_set = UnpairedBratsDataset(dataroot=args.dataroot, phase='train', transform=train_tf, source_domain=args.source_domain, target_domain=args.target_domain)
    val_set = PairedBratsValDataset(dataroot=args.dataroot, phase=args.eval_phase, transform=val_tf, source_domain=args.source_domain, target_domain=args.target_domain, pairing=args.eval_pairing)
    train_sampler = DistributedSampler(train_set, shuffle=True, drop_last=True) if world_size > 1 else None
    val_sampler = DistributedSampler(val_set, shuffle=False, drop_last=False) if world_size > 1 else None
    loader_kwargs = dict(num_workers=args.num_workers, pin_memory=True, persistent_workers=args.num_workers > 0, worker_init_fn=seed_worker)
    if args.num_workers > 0:
        loader_kwargs['prefetch_factor'] = args.prefetch_factor
    train_loader = DataLoader(train_set, batch_size=args.batch_size, shuffle=train_sampler is None, sampler=train_sampler, drop_last=True, **loader_kwargs)
    val_bs = args.val_batch_size if args.val_batch_size > 0 else args.batch_size * 4
    val_loader = DataLoader(val_set, batch_size=val_bs, shuffle=False, sampler=val_sampler, drop_last=False, **loader_kwargs)
    norm_layer = get_norm_layer(args.norm)
    netG_A = ResnetGenerator(input_nc=args.input_nc, output_nc=args.output_nc, ngf=args.ngf, norm_layer=norm_layer, use_dropout=False, n_blocks=args.netG_blocks).to(device)
    netG_B = ResnetGenerator(input_nc=args.output_nc, output_nc=args.input_nc, ngf=args.ngf, norm_layer=norm_layer, use_dropout=False, n_blocks=args.netG_blocks).to(device)
    netD_A = NLayerDiscriminator(input_nc=args.input_nc, ndf=args.ndf, n_layers=args.n_layers_D, norm_layer=norm_layer).to(device)
    netD_B = NLayerDiscriminator(input_nc=args.output_nc, ndf=args.ndf, n_layers=args.n_layers_D, norm_layer=norm_layer).to(device)
    afd = AdaptiveFrequencyDecomposition(input_nc=args.input_nc, embed_dim=args.afd_embed_dim, num_heads=args.afd_num_heads, high_min=args.afd_high_min, high_max=args.afd_high_max, low_min=args.afd_low_min, low_max=args.afd_low_max).to(device)
    with torch.no_grad():
        afd(torch.zeros(1, args.input_nc, args.afd_input_size, args.afd_input_size, device=device))
    if args.afd_resume:
        afd.load_state_dict(torch.load(args.afd_resume, map_location=device), strict=True)
    init_weights(netG_A, 'normal', 0.02)
    init_weights(netG_B, 'normal', 0.02)
    init_weights(netD_A, 'normal', 0.02)
    init_weights(netD_B, 'normal', 0.02)
    if args.channels_last:
        netG_A = netG_A.to(memory_format=torch.channels_last)
        netG_B = netG_B.to(memory_format=torch.channels_last)
        netD_A = netD_A.to(memory_format=torch.channels_last)
        netD_B = netD_B.to(memory_format=torch.channels_last)
    start_epoch = 1
    best_ssim = -1000000000.0
    best_psnr = -1000000000.0
    best_ssim_epoch = 0
    best_psnr_epoch = 0
    bad_epochs = 0
    ckpt = None
    if args.resume:
        ckpt_obj = torch.load(args.resume, map_location='cpu')
        if isinstance(ckpt_obj, dict) and 'G_A' in ckpt_obj:
            netG_A.load_state_dict(ckpt_obj['G_A'])
            if 'G_B' in ckpt_obj:
                netG_B.load_state_dict(ckpt_obj['G_B'])
            if 'D_A' in ckpt_obj:
                netD_A.load_state_dict(ckpt_obj['D_A'])
            if 'D_B' in ckpt_obj:
                netD_B.load_state_dict(ckpt_obj['D_B'])
            start_epoch = int(ckpt_obj.get('epoch', 0)) + 1
            best_ssim = float(ckpt_obj.get('best_ssim', best_ssim))
            best_psnr = float(ckpt_obj.get('best_psnr', best_psnr))
            best_ssim_epoch = int(ckpt_obj.get('best_ssim_epoch', best_ssim_epoch))
            best_psnr_epoch = int(ckpt_obj.get('best_psnr_epoch', best_psnr_epoch))
            bad_epochs = int(ckpt_obj.get('bad_epochs', bad_epochs))
            if is_main_process():
                print(f'[Resume] loaded old full checkpoint: {args.resume}, start_epoch={start_epoch}')
        else:
            netG_A.load_state_dict(ckpt_obj)
            start_epoch = 1
            best_ssim = -1000000000.0
            best_psnr = -1000000000.0
            best_ssim_epoch = 0
            best_psnr_epoch = 0
            bad_epochs = 0
            if is_main_process():
                print(f'[Resume] loaded G_A-only checkpoint: {args.resume}. Only G_A was loaded. G_B/D/optimizer are re-initialized. Training restarts from epoch 1.')
    if world_size > 1:
        netG_A = DDP(netG_A, device_ids=[local_rank], output_device=local_rank, find_unused_parameters=False)
        netG_B = DDP(netG_B, device_ids=[local_rank], output_device=local_rank, find_unused_parameters=False)
        netD_A = DDP(netD_A, device_ids=[local_rank], output_device=local_rank, find_unused_parameters=False)
        netD_B = DDP(netD_B, device_ids=[local_rank], output_device=local_rank, find_unused_parameters=False)
        afd = DDP(afd, device_ids=[local_rank], output_device=local_rank, find_unused_parameters=False)
    criterionGAN = GANLoss(args.gan_mode).to(device)
    criterionCycle = nn.L1Loss().to(device)
    criterionIdt = nn.L1Loss().to(device)
    optimizer_G = torch.optim.Adam(list(netG_A.parameters()) + list(netG_B.parameters()) + list(afd.parameters()), lr=args.lr, betas=(args.beta1, args.beta2))
    optimizer_D = torch.optim.Adam(list(netD_A.parameters()) + list(netD_B.parameters()), lr=args.lr, betas=(args.beta1, args.beta2))
    scaler = GradScaler(enabled=args.amp)
    fake_A_pool = ImagePool(args.pool_size)
    fake_B_pool = ImagePool(args.pool_size)
    total_epochs = args.n_epochs + args.n_epochs_decay
    global_step = 0
    for epoch in range(start_epoch, total_epochs + 1):
        if train_sampler is not None:
            train_sampler.set_epoch(epoch)
        if val_sampler is not None:
            val_sampler.set_epoch(epoch)
        lr = get_linear_decay_lr(epoch=epoch, base_lr=args.lr, n_epochs=args.n_epochs, n_epochs_decay=args.n_epochs_decay)
        set_optimizer_lr(optimizer_G, lr)
        set_optimizer_lr(optimizer_D, lr)
        netG_A.train()
        netG_B.train()
        netD_A.train()
        netD_B.train()
        afd.train()
        epoch_start = time.time()
        loss_acc = {'G': 0.0, 'G_A': 0.0, 'G_B': 0.0, 'cycle_A': 0.0, 'cycle_B': 0.0, 'idt_A': 0.0, 'idt_B': 0.0, 'FDDT': 0.0, 'FDDT_A2B': 0.0, 'FDDT_B2A': 0.0, 'FDDT_low': 0.0, 'FDDT_high': 0.0, 'D_A': 0.0, 'D_B': 0.0}
        n_iter = 0
        for batch in train_loader:
            real_A = batch['A'].to(device, non_blocking=True)
            real_B = batch['B'].to(device, non_blocking=True)
            if args.channels_last:
                real_A = real_A.contiguous(memory_format=torch.channels_last)
                real_B = real_B.contiguous(memory_format=torch.channels_last)
            set_requires_grad([netD_A, netD_B], False)
            optimizer_G.zero_grad(set_to_none=True)
            with autocast(enabled=args.amp):
                idt_B = netG_A(real_B)
                loss_idt_B = criterionIdt(idt_B, real_B) * args.lambda_B * args.lambda_identity
                idt_A = netG_B(real_A)
                loss_idt_A = criterionIdt(idt_A, real_A) * args.lambda_A * args.lambda_identity
                fake_B = netG_A(real_A)
                rec_A = netG_B(fake_B)
                fake_A = netG_B(real_B)
                rec_B = netG_A(fake_A)
                loss_G_A = criterionGAN(netD_B(fake_B), True)
                loss_G_B = criterionGAN(netD_A(fake_A), True)
                loss_cycle_A = criterionCycle(rec_A, real_A) * args.lambda_A
                loss_cycle_B = criterionCycle(rec_B, real_B) * args.lambda_B
                with autocast(enabled=False):
                    fddt_weight = get_fddt_weight(args, epoch)
                    loss_fddt_A2B_raw, loss_fddt_A2B_low, loss_fddt_A2B_high = adaptive_frequency_consistency_loss(generator=netG_A, afd=afd, source=real_A, full_translation=fake_B)[:3]
                    if args.fddt_bidir:
                        loss_fddt_B2A_raw, loss_fddt_B2A_low, loss_fddt_B2A_high = adaptive_frequency_consistency_loss(generator=netG_B, afd=afd, source=real_B, full_translation=fake_A)[:3]
                    else:
                        loss_fddt_B2A_raw = torch.zeros((), device=device)
                        loss_fddt_B2A_low = torch.zeros((), device=device)
                        loss_fddt_B2A_high = torch.zeros((), device=device)
                    loss_fddt_raw = loss_fddt_A2B_raw + loss_fddt_B2A_raw
                    loss_fddt = loss_fddt_raw * fddt_weight
                loss_G = loss_G_A + loss_G_B + loss_cycle_A + loss_cycle_B + loss_idt_A + loss_idt_B + loss_fddt
            scaler.scale(loss_G).backward()
            if args.grad_clip and args.grad_clip > 0:
                scaler.unscale_(optimizer_G)
                torch.nn.utils.clip_grad_norm_(list(netG_A.parameters()) + list(netG_B.parameters()) + list(afd.parameters()), args.grad_clip)
            scaler.step(optimizer_G)
            set_requires_grad([netD_A, netD_B], True)
            optimizer_D.zero_grad(set_to_none=True)
            fake_A_for_D = fake_A_pool.query(fake_A)
            fake_B_for_D = fake_B_pool.query(fake_B)
            with autocast(enabled=args.amp):
                loss_D_A_real = criterionGAN(netD_A(real_A), True)
                loss_D_A_fake = criterionGAN(netD_A(fake_A_for_D), False)
                loss_D_A = 0.5 * (loss_D_A_real + loss_D_A_fake)
                loss_D_B_real = criterionGAN(netD_B(real_B), True)
                loss_D_B_fake = criterionGAN(netD_B(fake_B_for_D), False)
                loss_D_B = 0.5 * (loss_D_B_real + loss_D_B_fake)
                loss_D = loss_D_A + loss_D_B
            scaler.scale(loss_D).backward()
            if args.grad_clip and args.grad_clip > 0:
                scaler.unscale_(optimizer_D)
                torch.nn.utils.clip_grad_norm_(list(netD_A.parameters()) + list(netD_B.parameters()), args.grad_clip)
            scaler.step(optimizer_D)
            scaler.update()
            loss_acc['G'] += float(loss_G.detach().item())
            loss_acc['G_A'] += float(loss_G_A.detach().item())
            loss_acc['G_B'] += float(loss_G_B.detach().item())
            loss_acc['cycle_A'] += float(loss_cycle_A.detach().item())
            loss_acc['cycle_B'] += float(loss_cycle_B.detach().item())
            loss_acc['idt_A'] += float(loss_idt_A.detach().item())
            loss_acc['idt_B'] += float(loss_idt_B.detach().item())
            loss_acc['FDDT'] += float(loss_fddt.detach().item())
            loss_acc['FDDT_A2B'] += float((loss_fddt_A2B_raw * fddt_weight).detach().item())
            loss_acc['FDDT_B2A'] += float((loss_fddt_B2A_raw * fddt_weight).detach().item())
            loss_acc['FDDT_low'] += float(((loss_fddt_A2B_low + loss_fddt_B2A_low) * fddt_weight).detach().item())
            loss_acc['FDDT_high'] += float(((loss_fddt_A2B_high + loss_fddt_B2A_high) * fddt_weight).detach().item())
            loss_acc['D_A'] += float(loss_D_A.detach().item())
            loss_acc['D_B'] += float(loss_D_B.detach().item())
            n_iter += 1
            global_step += 1
        val_start = time.time()
        case_rows, summary = validate_full_case_level(netG_A=netG_A, val_loader=val_loader, device=device, args=args)
        val_ssim = summary['case_ssim_mean']
        val_psnr = summary['case_psnr_mean']
        is_best_ssim = val_ssim > best_ssim + args.min_delta_ssim
        is_best_psnr = val_psnr > best_psnr + args.min_delta_psnr
        if is_best_ssim:
            best_ssim = val_ssim
            best_ssim_epoch = epoch
        if is_best_psnr:
            best_psnr = val_psnr
            best_psnr_epoch = epoch
        if args.early_metric == 'ssim':
            improved_for_early_stop = is_best_ssim
        elif args.early_metric == 'psnr':
            improved_for_early_stop = is_best_psnr
        else:
            improved_for_early_stop = is_best_ssim or is_best_psnr
        if improved_for_early_stop:
            bad_epochs = 0
        else:
            bad_epochs += 1
        if is_main_process():
            avg_losses = {k: v / max(1, n_iter) for k, v in loss_acc.items()}
            epoch_time = time.time() - epoch_start
            val_time = time.time() - val_start
            row = {'epoch': epoch, 'lr': lr, 'global_step': global_step, **avg_losses, 'val_case_ssim_mean': summary['case_ssim_mean'], 'val_case_ssim_std': summary['case_ssim_std'], 'val_case_psnr_mean': summary['case_psnr_mean'], 'val_case_psnr_std': summary['case_psnr_std'], 'val_n_cases': summary['n_cases'], 'val_n_slices': summary['n_slices'], 'best_ssim': best_ssim, 'best_ssim_epoch': best_ssim_epoch, 'best_psnr': best_psnr, 'best_psnr_epoch': best_psnr_epoch, 'is_best_ssim': int(is_best_ssim), 'is_best_psnr': int(is_best_psnr), 'bad_epochs': bad_epochs, 'epoch_sec': epoch_time, 'val_sec': val_time}
            append_csv(os.path.join(exp_dir, 'metrics.csv'), row)
            save_case_metrics_to_xlsx(xlsx_path=xlsx_path, epoch=epoch, case_rows=case_rows, summary=summary)
            save_G_A_only(os.path.join(exp_dir, f'epoch_{epoch:03d}_G_{args.source_domain}.pth'), netG_A)
            torch.save(unwrap_model(afd).state_dict(), os.path.join(exp_dir, f'epoch_{epoch:03d}_AFD.pth'))
            save_G_A_only(os.path.join(exp_dir, f'latest_G_{args.source_domain}.pth'), netG_A)
            torch.save(unwrap_model(afd).state_dict(), os.path.join(exp_dir, 'latest_AFD.pth'))
            if is_best_ssim:
                save_G_A_only(os.path.join(exp_dir, f'best_ssim_G_{args.source_domain}.pth'), netG_A)
                torch.save(unwrap_model(afd).state_dict(), os.path.join(exp_dir, 'best_ssim_AFD.pth'))
            if is_best_psnr:
                save_G_A_only(os.path.join(exp_dir, f'best_psnr_G_{args.source_domain}.pth'), netG_A)
                torch.save(unwrap_model(afd).state_dict(), os.path.join(exp_dir, 'best_psnr_AFD.pth'))
            print(f"[Epoch {epoch:03d}/{total_epochs:03d}] lr={lr:.2e} G={avg_losses['G']:.4f} FDDT={avg_losses['FDDT']:.4f} D_A={avg_losses['D_A']:.4f} D_B={avg_losses['D_B']:.4f} SSIM={summary['case_ssim_mean']:.5f}±{summary['case_ssim_std']:.5f} PSNR={summary['case_psnr_mean']:.3f}±{summary['case_psnr_std']:.3f} best_ssim={best_ssim:.5f}@{best_ssim_epoch} best_psnr={best_psnr:.3f}@{best_psnr_epoch} patience={bad_epochs}/{args.patience} train+val={epoch_time / 60:.1f}min val={val_time:.1f}s")
        stop_tensor = torch.tensor([1 if bad_epochs >= args.patience else 0], device=device, dtype=torch.int64)
        if dist.is_available() and dist.is_initialized():
            dist.broadcast(stop_tensor, src=0)
        if stop_tensor.item() == 1:
            if is_main_process():
                print(f'[EarlyStop] No improvement for {args.patience} epochs. Best SSIM={best_ssim:.5f}@{best_ssim_epoch}, Best PSNR={best_psnr:.3f}@{best_psnr_epoch}')
            break
    cleanup_ddp()
if __name__ == '__main__':
    main()
