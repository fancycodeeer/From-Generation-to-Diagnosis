import argparse
import os
import re
import time
from contextlib import nullcontext
from pathlib import Path
from typing import Dict, List, Tuple

import numpy as np
import torch
import torch.autograd as autograd
import torch.distributed as dist
import torch.nn.functional as F
from torch.cuda.amp import GradScaler
from torch.utils.data import DataLoader, DistributedSampler

from data import PairedBratsValDataset, UnpairedBratsDataset, build_transform
from models_syndiff import (
    SynDiffCoreBundle,
    apply_ema_to_generators,
    build_ema_helpers,
    build_syndiff_core_models,
    ddp_wrap_models,
    load_model_state,
    model_state_dict,
    restore_generators_from_ema,
    set_requires_grad,
    update_ema_helpers,
    unwrap_model,
)
from utils_ import (
    append_csv,
    is_main_process,
    psnr_per_image,
    seed_everything,
    seed_worker,
    ssim_per_image,
    tensor_to_01,
)


def str2bool(v):
    if isinstance(v, bool):
        return v
    v = str(v).lower()
    if v in ('yes', 'true', 't', '1', 'y'):
        return True
    if v in ('no', 'false', 'f', '0', 'n'):
        return False
    raise argparse.ArgumentTypeError('Boolean value expected.')


def parse_args():
    parser = argparse.ArgumentParser(
        'SynDiff-core, stable AMP, adapted to your project'
    )

    parser.add_argument('--dataroot', type=str, default='')
    parser.add_argument('--checkpoints_dir', type=str, default='')
    parser.add_argument('--name', type=str, default='t2_to_flair_syndiff')

    parser.add_argument('--mode', type=str, default='t2', choices=['t1', 't2'])
    parser.add_argument('--case_id_regex', type=str, default='')

    parser.add_argument('--input_nc', type=int, default=1)
    parser.add_argument('--output_nc', type=int, default=1)

    parser.add_argument('--image_size', type=int, default=256)
    parser.add_argument('--load_size', type=int, default=256)
    parser.add_argument('--crop_size', type=int, default=0)
    parser.add_argument('--no_flip', action='store_true', default=True)

    parser.add_argument('--num_workers', type=int, default=8)
    parser.add_argument('--prefetch_factor', type=int, default=4)

    parser.add_argument('--diff_base_ch', type=int, default=48)
    parser.add_argument('--ngf', type=int, default=64)
    parser.add_argument('--ndf', type=int, default=64)
    parser.add_argument('--resnet_blocks', type=int, default=6)
    parser.add_argument('--spectral_norm', type=str2bool, default=True)

    parser.add_argument('--nz', type=int, default=100)
    parser.add_argument('--t_emb_dim', type=int, default=256)

    parser.add_argument('--num_timesteps', type=int, default=4)
    parser.add_argument('--beta_min', type=float, default=0.1)
    parser.add_argument('--beta_max', type=float, default=20.0)
    parser.add_argument('--use_geometric', action='store_true', default=False)

    parser.add_argument('--batch_size', type=int, default=1)
    parser.add_argument('--val_batch_size', type=int, default=0)
    parser.add_argument('--num_epoch', type=int, default=500)

    # Paper-consistent SynDiff optimizer configuration
    parser.add_argument('--lr_g', type=float, default=1e-5)
    parser.add_argument('--lr_d', type=float, default=1e-5)
    parser.add_argument('--beta1', type=float, default=0.5)
    parser.add_argument('--beta2', type=float, default=0.9)
    parser.add_argument('--weight_decay', type=float, default=0.01)

    parser.add_argument('--no_lr_decay', action='store_true', default=False)

    parser.add_argument('--lambda_l1_loss', type=float, default=10.0)
    parser.add_argument('--r1_gamma', type=float, default=0.05)
    parser.add_argument('--lazy_reg', type=int, default=10)

    parser.add_argument('--grad_clip', type=float, default=1.0)

    parser.add_argument('--amp', type=str2bool, default=True)
    parser.add_argument('--safe_skip_nan', type=str2bool, default=True)
    parser.add_argument('--max_nan_batches', type=int, default=20)

    parser.add_argument('--use_ema', type=str2bool, default=True)
    parser.add_argument('--ema_decay', type=float, default=0.999)
    parser.add_argument('--eval_with_ema', type=str2bool, default=True)

    parser.add_argument('--seed', type=int, default=1024)
    parser.add_argument('--val_seed', type=int, default=1234)

    parser.add_argument('--print_freq', type=int, default=500)
    parser.add_argument('--save_content_every', type=int, default=10)
    parser.add_argument('--save_epoch_weights', action='store_true', default=False)

    parser.add_argument('--patience', type=int, default=20)
    parser.add_argument('--min_delta_ssim', type=float, default=1e-5)
    parser.add_argument('--min_delta_psnr', type=float, default=0.001)

    parser.add_argument('--resume', type=str, default='')

    parser.add_argument('--local_rank', type=int, default=-1)
    parser.add_argument('--local-rank', type=int, default=-1)

    return parser.parse_args()


def setup_ddp():
    if 'RANK' in os.environ and 'WORLD_SIZE' in os.environ:
        rank = int(os.environ['RANK'])
        world_size = int(os.environ['WORLD_SIZE'])
        local_rank = int(os.environ.get('LOCAL_RANK', 0))

        torch.cuda.set_device(local_rank)

        dist.init_process_group(
            backend='nccl',
            init_method='env://',
        )
    else:
        rank, world_size, local_rank = 0, 1, 0

    return rank, world_size, local_rank


def cleanup_ddp():
    if dist.is_available() and dist.is_initialized():
        dist.barrier()
        dist.destroy_process_group()


def ddp_all_gather_object(obj):
    if dist.is_available() and dist.is_initialized():
        gathered = [None for _ in range(dist.get_world_size())]
        dist.all_gather_object(gathered, obj)
        return gathered

    return [obj]


def amp_ctx(enabled: bool):
    enabled = bool(enabled and torch.cuda.is_available())

    if hasattr(torch, 'autocast'):
        return torch.autocast(
            device_type='cuda',
            dtype=torch.float16,
            enabled=enabled,
        )

    return nullcontext()


def is_finite_tensor(x: torch.Tensor) -> bool:
    return bool(torch.isfinite(x.detach()).all().item())


def all_finite(
    loss_dict: Dict[str, torch.Tensor],
) -> Tuple[bool, str]:
    for k, v in loss_dict.items():
        if not is_finite_tensor(v):
            return False, k

    return True, ''


def softplus_real(logits: torch.Tensor) -> torch.Tensor:
    return F.softplus(-logits.float()).mean()


def softplus_fake(logits: torch.Tensor) -> torch.Tensor:
    return F.softplus(logits.float()).mean()


def zero_optimizers(
    optimizers: Dict[str, torch.optim.Optimizer],
) -> None:
    for opt in optimizers.values():
        opt.zero_grad(set_to_none=True)


def var_func_vp(t, beta_min, beta_max):
    log_mean_coeff = (
        -0.25 * t ** 2 * (beta_max - beta_min)
        - 0.5 * t * beta_min
    )

    return 1.0 - torch.exp(2.0 * log_mean_coeff)


def var_func_geometric(t, beta_min, beta_max):
    return beta_min * (beta_max / beta_min) ** t


def extract(input_tensor, t, shape):
    out = torch.gather(input_tensor, 0, t)

    reshape = [shape[0]] + [1] * (len(shape) - 1)

    return out.reshape(*reshape)


def get_sigma_schedule(args, device):
    n_timestep = args.num_timesteps

    eps_small = 0.001

    t = (
        np.arange(
            0,
            n_timestep + 1,
            dtype=np.float64,
        )
        / n_timestep
    )

    t = torch.from_numpy(t) * (1.0 - eps_small) + eps_small

    if args.use_geometric:
        var = var_func_geometric(
            t,
            args.beta_min,
            args.beta_max,
        )
    else:
        var = var_func_vp(
            t,
            args.beta_min,
            args.beta_max,
        )

    alpha_bars = 1.0 - var

    betas = (
        1.0
        - alpha_bars[1:]
        / alpha_bars[:-1]
    )

    betas = torch.cat(
        (
            torch.tensor([1e-8]),
            betas,
        )
    ).to(
        device=device,
        dtype=torch.float32,
    )

    sigmas = betas.sqrt()

    a_s = torch.sqrt(
        1.0 - betas
    )

    return sigmas, a_s, betas


class Diffusion_Coefficients:

    def __init__(self, args, device):
        self.sigmas, self.a_s, _ = get_sigma_schedule(
            args,
            device=device,
        )

        a_s_cum = np.cumprod(
            self.a_s.detach().cpu()
        )

        self.a_s_cum = a_s_cum.to(
            device=device,
            dtype=torch.float32,
        )

        self.sigmas_cum = torch.sqrt(
            1.0 - self.a_s_cum ** 2
        )

        self.a_s_prev = self.a_s.clone()
        self.a_s_prev[-1] = 1.0


def q_sample(
    coeff: Diffusion_Coefficients,
    x_start,
    t,
    noise=None,
):
    if noise is None:
        noise = torch.randn_like(
            x_start
        )

    return (
        extract(
            coeff.a_s_cum,
            t,
            x_start.shape,
        )
        * x_start
        + extract(
            coeff.sigmas_cum,
            t,
            x_start.shape,
        )
        * noise
    )


def q_sample_pairs(
    coeff: Diffusion_Coefficients,
    x_start,
    t,
):
    noise = torch.randn_like(
        x_start
    )

    x_t = q_sample(
        coeff,
        x_start,
        t,
        noise=noise,
    )

    x_t_plus_one = (
        extract(
            coeff.a_s,
            t + 1,
            x_start.shape,
        )
        * x_t
        + extract(
            coeff.sigmas,
            t + 1,
            x_start.shape,
        )
        * noise
    )

    return x_t, x_t_plus_one


class Posterior_Coefficients:

    def __init__(self, args, device):
        _, _, betas = get_sigma_schedule(
            args,
            device=device,
        )

        self.betas = betas.float()[1:]

        self.alphas = (
            1.0
            - self.betas
        )

        self.alphas_cumprod = torch.cumprod(
            self.alphas,
            dim=0,
        )

        self.alphas_cumprod_prev = torch.cat(
            [
                torch.tensor(
                    [1.0],
                    dtype=torch.float32,
                    device=device,
                ),
                self.alphas_cumprod[:-1],
            ],
            dim=0,
        )

        self.posterior_variance = (
            self.betas
            * (
                1.0
                - self.alphas_cumprod_prev
            )
            / (
                1.0
                - self.alphas_cumprod
            )
        )

        self.posterior_mean_coef1 = (
            self.betas
            * torch.sqrt(
                self.alphas_cumprod_prev
            )
            / (
                1.0
                - self.alphas_cumprod
            )
        )

        self.posterior_mean_coef2 = (
            (
                1.0
                - self.alphas_cumprod_prev
            )
            * torch.sqrt(
                self.alphas
            )
            / (
                1.0
                - self.alphas_cumprod
            )
        )

        self.posterior_log_variance_clipped = torch.log(
            self.posterior_variance.clamp(
                min=1e-20
            )
        )


def sample_posterior(
    coefficients: Posterior_Coefficients,
    x_0,
    x_t,
    t,
):
    x_0 = x_0.float()
    x_t = x_t.float()

    mean = (
        extract(
            coefficients.posterior_mean_coef1,
            t,
            x_t.shape,
        )
        * x_0
        + extract(
            coefficients.posterior_mean_coef2,
            t,
            x_t.shape,
        )
        * x_t
    )

    log_var = extract(
        coefficients.posterior_log_variance_clipped,
        t,
        x_t.shape,
    )

    noise = torch.randn_like(
        x_t
    )

    nonzero_mask = (
        1.0
        - (t == 0).float()
    )

    return (
        mean
        + nonzero_mask[
            :,
            None,
            None,
            None,
        ]
        * torch.exp(
            0.5 * log_var
        )
        * noise
    )


@torch.no_grad()
def sample_from_model(
    coefficients,
    generator,
    n_time,
    x_init,
    args,
):
    x = x_init[:, [0], ...].float()
    source = x_init[:, [1], ...].float()

    generator.eval()

    for i in reversed(
        range(n_time)
    ):
        t = torch.full(
            (x.size(0),),
            i,
            dtype=torch.int64,
            device=x.device,
        )

        z = torch.randn(
            x.size(0),
            args.nz,
            device=x.device,
        )

        with amp_ctx(args.amp):
            x_0 = generator(
                torch.cat(
                    (
                        x,
                        source,
                    ),
                    dim=1,
                ),
                t,
                z,
            )

        x = sample_posterior(
            coefficients,
            x_0[:, [0], ...],
            x,
            t,
        )

        x = (
            x.detach()
            .clamp(
                -1.0,
                1.0,
            )
        )

    return x


def case_id_from_any_path_or_key(
    x: str,
    case_id_regex: str = '',
) -> str:
    stem = Path(
        str(x)
    ).stem

    if case_id_regex:
        match = re.search(
            case_id_regex,
            stem,
        )

        if match is None:
            raise ValueError(
                'case_id_regex did not match '
                f'evaluation file: {stem}'
            )

        return (
            match.group(1)
            if match.groups()
            else match.group(0)
        )

    if '_' in stem:
        parts = stem.split('_')

        if parts[0].startswith(
            'IXI'
        ):
            return parts[0]

        if (
            len(parts) >= 2
            and parts[0].startswith(
                'BraTS'
            )
        ):
            return (
                f'{parts[0]}_'
                f'{parts[1]}'
            )

        return (
            '_'.join(parts[:2])
            if len(parts) >= 2
            else stem
        )

    if '-' in stem:
        parts = stem.split('-')

        if (
            len(parts) >= 4
            and parts[0] == 'BraTS'
        ):
            return '-'.join(
                parts[:4]
            )

        return (
            '-'.join(
                parts[:-2]
            )
            if len(parts) >= 3
            else stem
        )

    return stem


def slice_id_from_any_path_or_key(
    x: str,
) -> str:
    return Path(
        str(x)
    ).stem


def compute_case_metrics_from_slice_entries(
    slice_entries: List[Dict],
) -> Tuple[List[Dict], Dict]:

    unique = {}

    for item in slice_entries:
        unique.setdefault(
            item['slice_id'],
            item,
        )

    grouped: Dict[
        str,
        List[Dict],
    ] = {}

    for item in unique.values():
        grouped.setdefault(
            item['case_id'],
            [],
        ).append(item)

    case_rows = []

    for case_id in sorted(
        grouped.keys()
    ):
        items = grouped[
            case_id
        ]

        ssim_arr = np.asarray(
            [
                x['ssim']
                for x in items
            ],
            dtype=np.float64,
        )

        psnr_arr = np.asarray(
            [
                x['psnr']
                for x in items
            ],
            dtype=np.float64,
        )

        case_rows.append(
            {
                'case_id': case_id,
                'n_slices': int(
                    len(items)
                ),
                'ssim_mean': float(
                    ssim_arr.mean()
                ),
                'ssim_std_slice': (
                    float(
                        ssim_arr.std(
                            ddof=1
                        )
                    )
                    if len(
                        ssim_arr
                    ) > 1
                    else 0.0
                ),
                'psnr_mean': float(
                    psnr_arr.mean()
                ),
                'psnr_std_slice': (
                    float(
                        psnr_arr.std(
                            ddof=1
                        )
                    )
                    if len(
                        psnr_arr
                    ) > 1
                    else 0.0
                ),
            }
        )

    case_ssim = np.asarray(
        [
            x['ssim_mean']
            for x in case_rows
        ],
        dtype=np.float64,
    )

    case_psnr = np.asarray(
        [
            x['psnr_mean']
            for x in case_rows
        ],
        dtype=np.float64,
    )

    summary = {
        'n_cases': int(
            len(case_rows)
        ),
        'n_slices': int(
            len(unique)
        ),
        'case_ssim_mean': (
            float(
                case_ssim.mean()
            )
            if len(case_ssim)
            else 0.0
        ),
        'case_ssim_std': (
            float(
                case_ssim.std(
                    ddof=1
                )
            )
            if len(case_ssim) > 1
            else 0.0
        ),
        'case_psnr_mean': (
            float(
                case_psnr.mean()
            )
            if len(case_psnr)
            else 0.0
        ),
        'case_psnr_std': (
            float(
                case_psnr.std(
                    ddof=1
                )
            )
            if len(case_psnr) > 1
            else 0.0
        ),
    }

    return (
        case_rows,
        summary,
    )


def save_case_metrics_to_xlsx(
    xlsx_path: str,
    epoch: int,
    case_rows: List[Dict],
    summary: Dict,
):
    try:
        from openpyxl import (
            Workbook,
            load_workbook,
        )
        from openpyxl.styles import (
            Alignment,
            Font,
        )
    except Exception:
        return

    xlsx_path = Path(
        xlsx_path
    )

    xlsx_path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    wb = (
        load_workbook(
            str(xlsx_path)
        )
        if xlsx_path.exists()
        else Workbook()
    )

    if (
        wb.active.title
        == 'Sheet'
    ):
        wb.active.title = (
            'case_level'
        )

    if (
        'case_level'
        not in wb.sheetnames
    ):
        wb.create_sheet(
            'case_level'
        )

    if (
        'epoch_summary'
        not in wb.sheetnames
    ):
        wb.create_sheet(
            'epoch_summary'
        )

    ws_case = wb[
        'case_level'
    ]

    ws_sum = wb[
        'epoch_summary'
    ]

    case_header = [
        'epoch',
        'case_id',
        'n_slices',
        'ssim_mean',
        'ssim_std_slice',
        'psnr_mean',
        'psnr_std_slice',
    ]

    sum_header = [
        'epoch',
        'n_cases',
        'n_slices',
        'case_ssim_mean',
        'case_ssim_std',
        'case_psnr_mean',
        'case_psnr_std',
    ]

    if (
        ws_case.max_row == 1
        and ws_case.cell(
            1,
            1,
        ).value is None
    ):
        ws_case.append(
            case_header
        )

    if (
        ws_sum.max_row == 1
        and ws_sum.cell(
            1,
            1,
        ).value is None
    ):
        ws_sum.append(
            sum_header
        )

    def remove_epoch_rows(ws):
        rows = [
            r
            for r in range(
                2,
                ws.max_row + 1,
            )
            if ws.cell(
                r,
                1,
            ).value == epoch
        ]

        for r in reversed(
            rows
        ):
            ws.delete_rows(
                r,
                1,
            )

    remove_epoch_rows(
        ws_case
    )
    remove_epoch_rows(
        ws_sum
    )

    for row in case_rows:
        ws_case.append(
            [
                epoch,
                row['case_id'],
                row['n_slices'],
                row['ssim_mean'],
                row['ssim_std_slice'],
                row['psnr_mean'],
                row['psnr_std_slice'],
            ]
        )

    ws_sum.append(
        [
            epoch,
            summary['n_cases'],
            summary['n_slices'],
            summary[
                'case_ssim_mean'
            ],
            summary[
                'case_ssim_std'
            ],
            summary[
                'case_psnr_mean'
            ],
            summary[
                'case_psnr_std'
            ],
        ]
    )

    for ws in [
        ws_case,
        ws_sum,
    ]:
        for cell in ws[1]:
            cell.font = Font(
                bold=True
            )

            cell.alignment = Alignment(
                horizontal='center'
            )

        ws.freeze_panes = 'A2'

    wb.save(
        str(xlsx_path)
    )


@torch.inference_mode()
def validate_A_to_B_case_level(
    models: SynDiffCoreBundle,
    val_loader,
    device,
    args,
    pos_coeff,
    ema_helpers=None,
):
    if (
        args.eval_with_ema
        and ema_helpers
    ):
        apply_ema_to_generators(
            ema_helpers,
            models,
        )

    models.gen_diffusive_2.eval()

    rank = (
        dist.get_rank()
        if (
            dist.is_available()
            and dist.is_initialized()
        )
        else 0
    )

    cpu_state = (
        torch.get_rng_state()
    )

    cuda_state = (
        torch.cuda.get_rng_state(
            device
        )
        if torch.cuda.is_available()
        else None
    )

    torch.manual_seed(
        args.val_seed + rank
    )

    if torch.cuda.is_available():
        torch.cuda.manual_seed(
            args.val_seed + rank
        )

    local_entries = []

    for batch in val_loader:
        real_A = batch[
            'A'
        ].to(
            device,
            non_blocking=True,
        )

        real_B = batch[
            'B'
        ].to(
            device,
            non_blocking=True,
        )

        noise = torch.randn_like(
            real_B
        )

        x_init = torch.cat(
            (
                noise,
                real_A,
            ),
            dim=1,
        )

        fake_B = sample_from_model(
            coefficients=pos_coeff,
            generator=models.gen_diffusive_2,
            n_time=args.num_timesteps,
            x_init=x_init,
            args=args,
        )

        fake01 = tensor_to_01(
            fake_B
        )

        realB01 = tensor_to_01(
            real_B
        )

        ssim_vals = (
            ssim_per_image(
                fake01,
                realB01,
            )
            .detach()
            .float()
            .cpu()
            .numpy()
        )

        psnr_vals = (
            psnr_per_image(
                fake01,
                realB01,
            )
            .detach()
            .float()
            .cpu()
            .numpy()
        )

        ids = batch.get(
            'key',
            batch.get(
                'A_path',
                [
                    str(i)
                    for i in range(
                        len(
                            ssim_vals
                        )
                    )
                ],
            ),
        )

        for i in range(
            len(ssim_vals)
        ):
            sid = ids[i]

            local_entries.append(
                {
                    'case_id': (
                        case_id_from_any_path_or_key(
                            sid,
                            args.case_id_regex,
                        )
                    ),
                    'slice_id': (
                        slice_id_from_any_path_or_key(
                            sid
                        )
                    ),
                    'ssim': float(
                        ssim_vals[i]
                    ),
                    'psnr': float(
                        psnr_vals[i]
                    ),
                }
            )

    torch.set_rng_state(
        cpu_state
    )

    if cuda_state is not None:
        torch.cuda.set_rng_state(
            cuda_state,
            device,
        )

    gathered = (
        ddp_all_gather_object(
            local_entries
        )
    )

    all_entries = []

    for part in gathered:
        all_entries.extend(
            part
        )

    case_rows, summary = (
        compute_case_metrics_from_slice_entries(
            all_entries
        )
    )

    models.gen_diffusive_2.train()

    if (
        args.eval_with_ema
        and ema_helpers
    ):
        restore_generators_from_ema(
            ema_helpers,
            models,
        )

    return (
        case_rows,
        summary,
    )


def save_full_checkpoint(
    path: str,
    epoch: int,
    global_step: int,
    models: SynDiffCoreBundle,
    optimizers: Dict,
    schedulers: Dict,
    scaler: GradScaler,
    args,
    best: Dict,
    ema_helpers: Dict,
):
    os.makedirs(
        os.path.dirname(path),
        exist_ok=True,
    )

    torch.save(
        {
            'epoch': epoch,
            'global_step': global_step,
            'args': vars(args),
            'best': best,
            'gen_diffusive_1': (
                model_state_dict(
                    models.gen_diffusive_1
                )
            ),
            'gen_diffusive_2': (
                model_state_dict(
                    models.gen_diffusive_2
                )
            ),
            'gen_non_diffusive_1to2': (
                model_state_dict(
                    models.gen_non_diffusive_1to2
                )
            ),
            'gen_non_diffusive_2to1': (
                model_state_dict(
                    models.gen_non_diffusive_2to1
                )
            ),
            'disc_diffusive_1': (
                model_state_dict(
                    models.disc_diffusive_1
                )
            ),
            'disc_diffusive_2': (
                model_state_dict(
                    models.disc_diffusive_2
                )
            ),
            'disc_cycle_1': (
                model_state_dict(
                    models.disc_cycle_1
                )
            ),
            'disc_cycle_2': (
                model_state_dict(
                    models.disc_cycle_2
                )
            ),
            'optimizers': {
                k: v.state_dict()
                for k, v
                in optimizers.items()
            },
            'schedulers': {
                k: v.state_dict()
                for k, v
                in schedulers.items()
            },
            'scaler': (
                scaler.state_dict()
            ),
            'ema': (
                {
                    k: v.state_dict()
                    for k, v
                    in ema_helpers.items()
                }
                if ema_helpers
                else {}
            ),
        },
        path,
    )


def load_full_checkpoint(
    path: str,
    models,
    optimizers,
    schedulers,
    scaler,
    ema_helpers,
    device,
):
    ckpt = torch.load(
        path,
        map_location=device,
    )

    load_model_state(
        models.gen_diffusive_1,
        ckpt[
            'gen_diffusive_1'
        ],
    )

    load_model_state(
        models.gen_diffusive_2,
        ckpt[
            'gen_diffusive_2'
        ],
    )

    load_model_state(
        models.gen_non_diffusive_1to2,
        ckpt[
            'gen_non_diffusive_1to2'
        ],
    )

    load_model_state(
        models.gen_non_diffusive_2to1,
        ckpt[
            'gen_non_diffusive_2to1'
        ],
    )

    load_model_state(
        models.disc_diffusive_1,
        ckpt[
            'disc_diffusive_1'
        ],
    )

    load_model_state(
        models.disc_diffusive_2,
        ckpt[
            'disc_diffusive_2'
        ],
    )

    load_model_state(
        models.disc_cycle_1,
        ckpt[
            'disc_cycle_1'
        ],
    )

    load_model_state(
        models.disc_cycle_2,
        ckpt[
            'disc_cycle_2'
        ],
    )

    for k, opt in optimizers.items():
        if (
            k
            in ckpt.get(
                'optimizers',
                {},
            )
        ):
            opt.load_state_dict(
                ckpt[
                    'optimizers'
                ][k]
            )

    for k, sch in schedulers.items():
        if (
            k
            in ckpt.get(
                'schedulers',
                {},
            )
        ):
            sch.load_state_dict(
                ckpt[
                    'schedulers'
                ][k]
            )

    if 'scaler' in ckpt:
        scaler.load_state_dict(
            ckpt[
                'scaler'
            ]
        )

    if (
        ema_helpers
        and 'ema' in ckpt
    ):
        for k, helper in (
            ema_helpers.items()
        ):
            helper.load_state_dict(
                ckpt[
                    'ema'
                ].get(k),
                device=device,
            )

    return ckpt


def save_A_to_B_weights(
    path: str,
    models: SynDiffCoreBundle,
):
    os.makedirs(
        os.path.dirname(path),
        exist_ok=True,
    )

    torch.save(
        model_state_dict(
            models.gen_diffusive_2
        ),
        path,
    )


def main():
    args = parse_args()

    rank, world_size, local_rank = (
        setup_ddp()
    )

    device = torch.device(
        (
            f'cuda:{local_rank}'
            if torch.cuda.is_available()
            else 'cpu'
        )
    )

    if (
        world_size > 1
        and args.safe_skip_nan
    ):
        if is_main_process():
            print(
                '[DDP] safe_skip_nan is disabled '
                'under DDP to avoid reducer-state mismatch.'
            )

        args.safe_skip_nan = False

    seed_everything(
        args.seed + rank
    )

    torch.backends.cudnn.benchmark = True

    try:
        torch.set_float32_matmul_precision(
            'high'
        )
    except Exception:
        pass

    exp_dir = os.path.join(
        args.checkpoints_dir,
        args.name,
    )

    xlsx_path = os.path.join(
        exp_dir,
        'case_level_metrics.xlsx',
    )

    metrics_csv = os.path.join(
        exp_dir,
        'metrics.csv',
    )

    if is_main_process():
        os.makedirs(
            exp_dir,
            exist_ok=True,
        )

        print(
            f'[DDP] world_size={world_size}'
        )

        print(
            f'[AMP] enabled={args.amp}, '
            f'GradScaler=True, '
            f'safe_skip_nan={args.safe_skip_nan}'
        )

        print(
            f'[Data] train A: '
            f'{args.dataroot}/train/{args.mode}'
        )

        print(
            f'[Data] val A:   '
            f'{args.dataroot}/val/{args.mode}'
        )

        print(
            f'[SynDiff-core] '
            f'T={args.num_timesteps}, '
            f'beta_min={args.beta_min}, '
            f'beta_max={args.beta_max}, '
            f'geometric={args.use_geometric}'
        )

        print(
            f'[Optimizer] AdamW | '
            f'lr_g={args.lr_g:.2e} | '
            f'lr_d={args.lr_d:.2e} | '
            f'betas=({args.beta1}, {args.beta2}) | '
            f'weight_decay={args.weight_decay}'
        )

        print(
            f'[Model] '
            f'diff_base_ch={args.diff_base_ch}, '
            f'ngf={args.ngf}, '
            f'ndf={args.ndf}, '
            f'spectral_norm={args.spectral_norm}'
        )

        print(
            f'[Save] {exp_dir}'
        )

    train_tf = build_transform(
        'train',
        args.input_nc,
        args.load_size,
        args.crop_size,
        no_flip=True,
    )

    val_tf = build_transform(
        'val',
        args.input_nc,
        args.load_size,
        args.crop_size,
        no_flip=True,
    )

    train_set = UnpairedBratsDataset(
        args.dataroot,
        'train',
        train_tf,
        args.mode,
    )

    val_set = PairedBratsValDataset(
        args.dataroot,
        'val',
        val_tf,
        args.mode,
    )

    train_sampler = (
        DistributedSampler(
            train_set,
            shuffle=True,
            drop_last=True,
        )
        if world_size > 1
        else None
    )

    val_sampler = (
        DistributedSampler(
            val_set,
            shuffle=False,
            drop_last=False,
        )
        if world_size > 1
        else None
    )

    loader_kwargs = dict(
        num_workers=args.num_workers,
        pin_memory=True,
        persistent_workers=(
            args.num_workers > 0
        ),
        worker_init_fn=seed_worker,
    )

    if args.num_workers > 0:
        loader_kwargs[
            'prefetch_factor'
        ] = args.prefetch_factor

    train_loader = DataLoader(
        train_set,
        batch_size=args.batch_size,
        shuffle=(
            train_sampler is None
        ),
        sampler=train_sampler,
        drop_last=True,
        **loader_kwargs,
    )

    val_bs = (
        args.val_batch_size
        if args.val_batch_size > 0
        else args.batch_size * 4
    )

    val_loader = DataLoader(
        val_set,
        batch_size=val_bs,
        shuffle=False,
        sampler=val_sampler,
        drop_last=False,
        **loader_kwargs,
    )

    coeff = Diffusion_Coefficients(
        args,
        device,
    )

    pos_coeff = Posterior_Coefficients(
        args,
        device,
    )

    models = build_syndiff_core_models(
        args,
        device,
    )

    # ---------------------------------------------------------
    # AdamW optimizers — paper-consistent SynDiff configuration
    # ---------------------------------------------------------
    optimizers = {
        'gen_diffusive_1': torch.optim.AdamW(
            models.gen_diffusive_1.parameters(),
            lr=args.lr_g,
            betas=(
                args.beta1,
                args.beta2,
            ),
            weight_decay=args.weight_decay,
        ),
        'gen_diffusive_2': torch.optim.AdamW(
            models.gen_diffusive_2.parameters(),
            lr=args.lr_g,
            betas=(
                args.beta1,
                args.beta2,
            ),
            weight_decay=args.weight_decay,
        ),
        'gen_non_diffusive_1to2': torch.optim.AdamW(
            models.gen_non_diffusive_1to2.parameters(),
            lr=args.lr_g,
            betas=(
                args.beta1,
                args.beta2,
            ),
            weight_decay=args.weight_decay,
        ),
        'gen_non_diffusive_2to1': torch.optim.AdamW(
            models.gen_non_diffusive_2to1.parameters(),
            lr=args.lr_g,
            betas=(
                args.beta1,
                args.beta2,
            ),
            weight_decay=args.weight_decay,
        ),
        'disc_diffusive_1': torch.optim.AdamW(
            models.disc_diffusive_1.parameters(),
            lr=args.lr_d,
            betas=(
                args.beta1,
                args.beta2,
            ),
            weight_decay=args.weight_decay,
        ),
        'disc_diffusive_2': torch.optim.AdamW(
            models.disc_diffusive_2.parameters(),
            lr=args.lr_d,
            betas=(
                args.beta1,
                args.beta2,
            ),
            weight_decay=args.weight_decay,
        ),
        'disc_cycle_1': torch.optim.AdamW(
            models.disc_cycle_1.parameters(),
            lr=args.lr_d,
            betas=(
                args.beta1,
                args.beta2,
            ),
            weight_decay=args.weight_decay,
        ),
        'disc_cycle_2': torch.optim.AdamW(
            models.disc_cycle_2.parameters(),
            lr=args.lr_d,
            betas=(
                args.beta1,
                args.beta2,
            ),
            weight_decay=args.weight_decay,
        ),
    }

    schedulers = {}

    if not args.no_lr_decay:
        for k, opt in optimizers.items():
            schedulers[
                k
            ] = torch.optim.lr_scheduler.CosineAnnealingLR(
                opt,
                T_max=args.num_epoch,
                eta_min=1e-5,
            )

    scaler = GradScaler(
        enabled=(
            args.amp
            and device.type == 'cuda'
        )
    )

    ema_helpers = (
        build_ema_helpers(
            args,
            models,
        )
    )

    start_epoch = 0
    global_step = 0

    best = {
        'best_ssim': -1e9,
        'best_psnr': -1e9,
        'best_ssim_epoch': -1,
        'best_psnr_epoch': -1,
        'bad_epochs': 0,
    }

    if args.resume:
        ckpt = load_full_checkpoint(
            args.resume,
            models,
            optimizers,
            schedulers,
            scaler,
            ema_helpers,
            device,
        )

        start_epoch = (
            int(
                ckpt['epoch']
            )
            + 1
        )

        global_step = int(
            ckpt.get(
                'global_step',
                0,
            )
        )

        best.update(
            ckpt.get(
                'best',
                {},
            )
        )

        if is_main_process():
            print(
                f'[Resume] loaded '
                f'{args.resume}, '
                f'start_epoch={start_epoch}'
            )

    models = ddp_wrap_models(
        models,
        local_rank=local_rank,
        find_unused_parameters=False,
    )

    nan_batches = 0

    try:
        for epoch in range(
            start_epoch,
            args.num_epoch + 1,
        ):
            if train_sampler is not None:
                train_sampler.set_epoch(
                    epoch
                )

            epoch_start = time.time()

            for net in [
                models.gen_diffusive_1,
                models.gen_diffusive_2,
                models.gen_non_diffusive_1to2,
                models.gen_non_diffusive_2to1,
                models.disc_diffusive_1,
                models.disc_diffusive_2,
                models.disc_cycle_1,
                models.disc_cycle_2,
            ]:
                net.train()

            loss_acc = {
                'G_sum': 0.0,
                'G_cycle': 0.0,
                'G_L1': 0.0,
                'G_adv': 0.0,
                'G_cycle_adv': 0.0,
                'D_diff': 0.0,
                'D_cycle': 0.0,
                'R1': 0.0,
            }

            n_iter = 0

            for batch in train_loader:
                real_data1 = batch[
                    'A'
                ].to(
                    device,
                    non_blocking=True,
                )

                real_data2 = batch[
                    'B'
                ].to(
                    device,
                    non_blocking=True,
                )

                bsz = real_data1.size(
                    0
                )

                r1_value = torch.zeros(
                    (),
                    device=device,
                    dtype=torch.float32,
                )

                set_requires_grad(
                    [
                        models.disc_diffusive_1,
                        models.disc_diffusive_2,
                        models.disc_cycle_1,
                        models.disc_cycle_2,
                    ],
                    True,
                )

                for k in [
                    'disc_diffusive_1',
                    'disc_diffusive_2',
                    'disc_cycle_1',
                    'disc_cycle_2',
                ]:
                    optimizers[
                        k
                    ].zero_grad(
                        set_to_none=True
                    )

                t1 = torch.randint(
                    0,
                    args.num_timesteps,
                    (bsz,),
                    device=device,
                )

                t2 = torch.randint(
                    0,
                    args.num_timesteps,
                    (bsz,),
                    device=device,
                )

                x1_t, x1_tp1 = (
                    q_sample_pairs(
                        coeff,
                        real_data1,
                        t1,
                    )
                )

                x2_t, x2_tp1 = (
                    q_sample_pairs(
                        coeff,
                        real_data2,
                        t2,
                    )
                )

                with torch.no_grad():
                    z1 = torch.randn(
                        bsz,
                        args.nz,
                        device=device,
                    )

                    z2 = torch.randn(
                        bsz,
                        args.nz,
                        device=device,
                    )

                    G_2to1 = unwrap_model(
                        models.gen_non_diffusive_2to1
                    )

                    G_1to2 = unwrap_model(
                        models.gen_non_diffusive_1to2
                    )

                    G_diff_1 = unwrap_model(
                        models.gen_diffusive_1
                    )

                    G_diff_2 = unwrap_model(
                        models.gen_diffusive_2
                    )

                    x1_0_predict = (
                        G_2to1(
                            real_data2
                        )
                    )

                    x2_0_predict = (
                        G_1to2(
                            real_data1
                        )
                    )

                    with amp_ctx(
                        args.amp
                    ):
                        x1_0_predict_diff = G_diff_1(
                            torch.cat(
                                (
                                    x1_tp1.detach(),
                                    x2_0_predict.detach(),
                                ),
                                dim=1,
                            ),
                            t1,
                            z1,
                        )

                        x2_0_predict_diff = G_diff_2(
                            torch.cat(
                                (
                                    x2_tp1.detach(),
                                    x1_0_predict.detach(),
                                ),
                                dim=1,
                            ),
                            t2,
                            z2,
                        )

                    x1_pos_sample = sample_posterior(
                        pos_coeff,
                        x1_0_predict_diff[
                            :,
                            [0],
                            ...,
                        ],
                        x1_tp1,
                        t1,
                    )

                    x2_pos_sample = sample_posterior(
                        pos_coeff,
                        x2_0_predict_diff[
                            :,
                            [0],
                            ...,
                        ],
                        x2_tp1,
                        t2,
                    )

                need_r1 = (
                    args.r1_gamma > 0
                    and (
                        args.lazy_reg is None
                        or global_step
                        % args.lazy_reg
                        == 0
                    )
                )

                x1_t_for_d = (
                    x1_t.detach()
                    .float()
                )

                x2_t_for_d = (
                    x2_t.detach()
                    .float()
                )

                if need_r1:
                    x1_t_for_d.requires_grad_(
                        True
                    )

                    x2_t_for_d.requires_grad_(
                        True
                    )

                with amp_ctx(
                    args.amp
                    and not need_r1
                ):
                    D1_real = models.disc_diffusive_1(
                        x1_t_for_d,
                        t1,
                        x1_tp1.detach().float(),
                    ).view(-1)

                    D2_real = models.disc_diffusive_2(
                        x2_t_for_d,
                        t2,
                        x2_tp1.detach().float(),
                    ).view(-1)

                with amp_ctx(
                    args.amp
                ):
                    D1_fake = models.disc_diffusive_1(
                        x1_pos_sample.detach(),
                        t1,
                        x1_tp1.detach(),
                    ).view(-1)

                    D2_fake = models.disc_diffusive_2(
                        x2_pos_sample.detach(),
                        t2,
                        x2_tp1.detach(),
                    ).view(-1)

                errD_diff = (
                    softplus_real(
                        D1_real
                    )
                    + softplus_real(
                        D2_real
                    )
                    + softplus_fake(
                        D1_fake
                    )
                    + softplus_fake(
                        D2_fake
                    )
                )

                ok, bad_name = all_finite(
                    {
                        'D_diff': errD_diff
                    }
                )

                if not ok:
                    nan_batches += 1

                    zero_optimizers(
                        optimizers
                    )

                    if is_main_process():
                        print(
                            f'[SafeSkip] non-finite '
                            f'{bad_name} '
                            f'at step={global_step}, '
                            f'skipped={nan_batches}'
                        )

                    if (
                        not args.safe_skip_nan
                        or nan_batches
                        > args.max_nan_batches
                    ):
                        raise RuntimeError(
                            f'Non-finite '
                            f'{bad_name}'
                        )

                    continue

                scaler.scale(
                    errD_diff
                ).backward(
                    retain_graph=need_r1
                )

                if need_r1:
                    grad1 = autograd.grad(
                        outputs=D1_real.sum(),
                        inputs=x1_t_for_d,
                        create_graph=True,
                        retain_graph=True,
                        only_inputs=True,
                    )[0]

                    grad2 = autograd.grad(
                        outputs=D2_real.sum(),
                        inputs=x2_t_for_d,
                        create_graph=True,
                        retain_graph=True,
                        only_inputs=True,
                    )[0]

                    r1_loss = (
                        args.r1_gamma
                        * 0.5
                        * (
                            grad1.flatten(
                                1
                            )
                            .pow(2)
                            .sum(1)
                            .mean()
                            + grad2.flatten(
                                1
                            )
                            .pow(2)
                            .sum(1)
                            .mean()
                        )
                    )

                    ok, bad_name = all_finite(
                        {
                            'R1': r1_loss
                        }
                    )

                    if not ok:
                        nan_batches += 1

                        zero_optimizers(
                            optimizers
                        )

                        if is_main_process():
                            print(
                                f'[SafeSkip] non-finite '
                                f'{bad_name} '
                                f'at step={global_step}, '
                                f'skipped={nan_batches}'
                            )

                        if (
                            not args.safe_skip_nan
                            or nan_batches
                            > args.max_nan_batches
                        ):
                            raise RuntimeError(
                                f'Non-finite '
                                f'{bad_name}'
                            )

                        continue

                    scaler.scale(
                        r1_loss
                    ).backward()

                    r1_value = (
                        r1_loss.detach()
                    )

                with torch.no_grad():
                    G_2to1 = unwrap_model(
                        models.gen_non_diffusive_2to1
                    )

                    G_1to2 = unwrap_model(
                        models.gen_non_diffusive_1to2
                    )

                    x1_0_predict = G_2to1(
                        real_data2
                    )

                    x2_0_predict = G_1to2(
                        real_data1
                    )

                with amp_ctx(
                    args.amp
                ):
                    D_cycle1_real = (
                        models.disc_cycle_1(
                            real_data1
                        ).view(-1)
                    )

                    D_cycle2_real = (
                        models.disc_cycle_2(
                            real_data2
                        ).view(-1)
                    )

                    D_cycle1_fake = (
                        models.disc_cycle_1(
                            x1_0_predict.detach()
                        ).view(-1)
                    )

                    D_cycle2_fake = (
                        models.disc_cycle_2(
                            x2_0_predict.detach()
                        ).view(-1)
                    )

                    errD_cycle = (
                        softplus_real(
                            D_cycle1_real
                        )
                        + softplus_real(
                            D_cycle2_real
                        )
                        + softplus_fake(
                            D_cycle1_fake
                        )
                        + softplus_fake(
                            D_cycle2_fake
                        )
                    )

                ok, bad_name = all_finite(
                    {
                        'D_cycle': errD_cycle
                    }
                )

                if not ok:
                    nan_batches += 1

                    zero_optimizers(
                        optimizers
                    )

                    if is_main_process():
                        print(
                            f'[SafeSkip] non-finite '
                            f'{bad_name} '
                            f'at step={global_step}, '
                            f'skipped={nan_batches}'
                        )

                    if (
                        not args.safe_skip_nan
                        or nan_batches
                        > args.max_nan_batches
                    ):
                        raise RuntimeError(
                            f'Non-finite '
                            f'{bad_name}'
                        )

                    continue

                scaler.scale(
                    errD_cycle
                ).backward()

                for k in [
                    'disc_diffusive_1',
                    'disc_diffusive_2',
                    'disc_cycle_1',
                    'disc_cycle_2',
                ]:
                    scaler.unscale_(
                        optimizers[k]
                    )

                if args.grad_clip > 0:
                    torch.nn.utils.clip_grad_norm_(
                        models.disc_diffusive_1.parameters(),
                        args.grad_clip,
                    )

                    torch.nn.utils.clip_grad_norm_(
                        models.disc_diffusive_2.parameters(),
                        args.grad_clip,
                    )

                    torch.nn.utils.clip_grad_norm_(
                        models.disc_cycle_1.parameters(),
                        args.grad_clip,
                    )

                    torch.nn.utils.clip_grad_norm_(
                        models.disc_cycle_2.parameters(),
                        args.grad_clip,
                    )

                for k in [
                    'disc_diffusive_1',
                    'disc_diffusive_2',
                    'disc_cycle_1',
                    'disc_cycle_2',
                ]:
                    scaler.step(
                        optimizers[k]
                    )

                set_requires_grad(
                    [
                        models.disc_diffusive_1,
                        models.disc_diffusive_2,
                        models.disc_cycle_1,
                        models.disc_cycle_2,
                    ],
                    False,
                )

                for k in [
                    'gen_diffusive_1',
                    'gen_diffusive_2',
                    'gen_non_diffusive_1to2',
                    'gen_non_diffusive_2to1',
                ]:
                    optimizers[
                        k
                    ].zero_grad(
                        set_to_none=True
                    )

                t1 = torch.randint(
                    0,
                    args.num_timesteps,
                    (bsz,),
                    device=device,
                )

                t2 = torch.randint(
                    0,
                    args.num_timesteps,
                    (bsz,),
                    device=device,
                )

                x1_t, x1_tp1 = (
                    q_sample_pairs(
                        coeff,
                        real_data1,
                        t1,
                    )
                )

                x2_t, x2_tp1 = (
                    q_sample_pairs(
                        coeff,
                        real_data2,
                        t2,
                    )
                )

                z1 = torch.randn(
                    bsz,
                    args.nz,
                    device=device,
                )

                z2 = torch.randn(
                    bsz,
                    args.nz,
                    device=device,
                )

                D_diff_1 = unwrap_model(
                    models.disc_diffusive_1
                )

                D_diff_2 = unwrap_model(
                    models.disc_diffusive_2
                )

                D_cyc_1 = unwrap_model(
                    models.disc_cycle_1
                )

                D_cyc_2 = unwrap_model(
                    models.disc_cycle_2
                )

                with amp_ctx(
                    args.amp
                ):
                    x1_0_predict = (
                        models.gen_non_diffusive_2to1(
                            real_data2
                        )
                    )

                    x2_0_predict_cycle = (
                        models.gen_non_diffusive_1to2(
                            x1_0_predict
                        )
                    )

                    x2_0_predict = (
                        models.gen_non_diffusive_1to2(
                            real_data1
                        )
                    )

                    x1_0_predict_cycle = (
                        models.gen_non_diffusive_2to1(
                            x2_0_predict
                        )
                    )

                    x1_0_predict_diff = (
                        models.gen_diffusive_1(
                            torch.cat(
                                (
                                    x1_tp1.detach(),
                                    x2_0_predict,
                                ),
                                dim=1,
                            ),
                            t1,
                            z1,
                        )
                    )

                    x2_0_predict_diff = (
                        models.gen_diffusive_2(
                            torch.cat(
                                (
                                    x2_tp1.detach(),
                                    x1_0_predict,
                                ),
                                dim=1,
                            ),
                            t2,
                            z2,
                        )
                    )

                    x1_pos_sample = sample_posterior(
                        pos_coeff,
                        x1_0_predict_diff[
                            :,
                            [0],
                            ...,
                        ],
                        x1_tp1,
                        t1,
                    )

                    x2_pos_sample = sample_posterior(
                        pos_coeff,
                        x2_0_predict_diff[
                            :,
                            [0],
                            ...,
                        ],
                        x2_tp1,
                        t2,
                    )

                    output1 = D_diff_1(
                        x1_pos_sample,
                        t1,
                        x1_tp1.detach(),
                    ).view(-1)

                    output2 = D_diff_2(
                        x2_pos_sample,
                        t2,
                        x2_tp1.detach(),
                    ).view(-1)

                    errG_adv = (
                        softplus_real(
                            output1
                        )
                        + softplus_real(
                            output2
                        )
                    )

                    D_cycle1_fake = (
                        D_cyc_1(
                            x1_0_predict
                        ).view(-1)
                    )

                    D_cycle2_fake = (
                        D_cyc_2(
                            x2_0_predict
                        ).view(-1)
                    )

                    errG_cycle_adv = (
                        softplus_real(
                            D_cycle1_fake
                        )
                        + softplus_real(
                            D_cycle2_fake
                        )
                    )

                    errG_L1 = (
                        F.l1_loss(
                            x1_0_predict_diff[
                                :,
                                [0],
                                ...,
                            ].float(),
                            real_data1.float(),
                        )
                        + F.l1_loss(
                            x2_0_predict_diff[
                                :,
                                [0],
                                ...,
                            ].float(),
                            real_data2.float(),
                        )
                    )

                    errG_cycle = (
                        F.l1_loss(
                            x1_0_predict_cycle.float(),
                            real_data1.float(),
                        )
                        + F.l1_loss(
                            x2_0_predict_cycle.float(),
                            real_data2.float(),
                        )
                    )

                    errG = (
                        args.lambda_l1_loss
                        * errG_cycle
                        + errG_adv
                        + errG_cycle_adv
                        + args.lambda_l1_loss
                        * errG_L1
                    )

                ok, bad_name = all_finite(
                    {
                        'G_sum': errG,
                        'G_cycle': errG_cycle,
                        'G_L1': errG_L1,
                        'G_adv': errG_adv,
                        'G_cycle_adv': (
                            errG_cycle_adv
                        ),
                    }
                )

                if not ok:
                    nan_batches += 1

                    zero_optimizers(
                        optimizers
                    )

                    if is_main_process():
                        print(
                            f'[SafeSkip] non-finite '
                            f'{bad_name} '
                            f'at step={global_step}, '
                            f'skipped={nan_batches}'
                        )

                    if (
                        not args.safe_skip_nan
                        or nan_batches
                        > args.max_nan_batches
                    ):
                        raise RuntimeError(
                            f'Non-finite '
                            f'{bad_name}'
                        )

                    continue

                scaler.scale(
                    errG
                ).backward()

                for k in [
                    'gen_diffusive_1',
                    'gen_diffusive_2',
                    'gen_non_diffusive_1to2',
                    'gen_non_diffusive_2to1',
                ]:
                    scaler.unscale_(
                        optimizers[k]
                    )

                if args.grad_clip > 0:
                    torch.nn.utils.clip_grad_norm_(
                        models.gen_diffusive_1.parameters(),
                        args.grad_clip,
                    )

                    torch.nn.utils.clip_grad_norm_(
                        models.gen_diffusive_2.parameters(),
                        args.grad_clip,
                    )

                    torch.nn.utils.clip_grad_norm_(
                        models.gen_non_diffusive_1to2.parameters(),
                        args.grad_clip,
                    )

                    torch.nn.utils.clip_grad_norm_(
                        models.gen_non_diffusive_2to1.parameters(),
                        args.grad_clip,
                    )

                for k in [
                    'gen_diffusive_1',
                    'gen_diffusive_2',
                    'gen_non_diffusive_1to2',
                    'gen_non_diffusive_2to1',
                ]:
                    scaler.step(
                        optimizers[k]
                    )

                scaler.update()

                update_ema_helpers(
                    ema_helpers,
                    models,
                )

                nan_batches = 0
                global_step += 1
                n_iter += 1

                loss_acc[
                    'G_sum'
                ] += float(
                    errG.detach()
                )

                loss_acc[
                    'G_cycle'
                ] += float(
                    errG_cycle.detach()
                )

                loss_acc[
                    'G_L1'
                ] += float(
                    errG_L1.detach()
                )

                loss_acc[
                    'G_adv'
                ] += float(
                    errG_adv.detach()
                )

                loss_acc[
                    'G_cycle_adv'
                ] += float(
                    errG_cycle_adv.detach()
                )

                loss_acc[
                    'D_diff'
                ] += float(
                    errD_diff.detach()
                )

                loss_acc[
                    'D_cycle'
                ] += float(
                    errD_cycle.detach()
                )

                loss_acc[
                    'R1'
                ] += float(
                    r1_value.detach()
                )

                if (
                    is_main_process()
                    and global_step
                    % args.print_freq
                    == 0
                ):
                    print(
                        f'[Epoch '
                        f'{epoch:03d}/'
                        f'{args.num_epoch}] '
                        f'iter={n_iter:04d} '
                        f'G-Cycle='
                        f'{float(errG_cycle.detach()):.4f} '
                        f'G-L1='
                        f'{float(errG_L1.detach()):.4f} '
                        f'G-Adv='
                        f'{float(errG_adv.detach()):.4f} '
                        f'G-cycle-Adv='
                        f'{float(errG_cycle_adv.detach()):.4f} '
                        f'G-Sum='
                        f'{float(errG.detach()):.4f} '
                        f'D='
                        f'{float(errD_diff.detach()):.4f} '
                        f'D-cycle='
                        f'{float(errD_cycle.detach()):.4f} '
                        f'R1='
                        f'{float(r1_value.detach()):.4f} '
                        f'scale='
                        f'{scaler.get_scale():.1f}'
                    )

            for sch in (
                schedulers.values()
            ):
                sch.step()

            case_rows, summary = (
                validate_A_to_B_case_level(
                    models,
                    val_loader,
                    device,
                    args,
                    pos_coeff,
                    ema_helpers,
                )
            )

            if is_main_process():
                save_case_metrics_to_xlsx(
                    xlsx_path,
                    epoch,
                    case_rows,
                    summary,
                )

                avg = {
                    k: (
                        v
                        / max(
                            n_iter,
                            1,
                        )
                    )
                    for k, v
                    in loss_acc.items()
                }

                lr_g = (
                    optimizers[
                        'gen_diffusive_2'
                    ]
                    .param_groups[0][
                        'lr'
                    ]
                )

                lr_d = (
                    optimizers[
                        'disc_diffusive_2'
                    ]
                    .param_groups[0][
                        'lr'
                    ]
                )

                cur_ssim = (
                    summary[
                        'case_ssim_mean'
                    ]
                )

                cur_psnr = (
                    summary[
                        'case_psnr_mean'
                    ]
                )

                improved = False

                if (
                    cur_ssim
                    > best[
                        'best_ssim'
                    ]
                    + args.min_delta_ssim
                ):
                    best[
                        'best_ssim'
                    ] = cur_ssim

                    best[
                        'best_ssim_epoch'
                    ] = epoch

                    best[
                        'bad_epochs'
                    ] = 0

                    improved = True

                    best_ssim_path = os.path.join(
                        exp_dir,
                        'best_ssim_netG_A2B.pth',
                    )

                    if (
                        args.eval_with_ema
                        and ema_helpers
                    ):
                        apply_ema_to_generators(
                            ema_helpers,
                            models,
                        )

                        save_A_to_B_weights(
                            best_ssim_path,
                            models,
                        )

                        restore_generators_from_ema(
                            ema_helpers,
                            models,
                        )
                    else:
                        save_A_to_B_weights(
                            best_ssim_path,
                            models,
                        )

                if (
                    cur_psnr
                    > best[
                        'best_psnr'
                    ]
                    + args.min_delta_psnr
                ):
                    best[
                        'best_psnr'
                    ] = cur_psnr

                    best[
                        'best_psnr_epoch'
                    ] = epoch

                    best[
                        'bad_epochs'
                    ] = 0

                    improved = True

                    best_psnr_path = os.path.join(
                        exp_dir,
                        'best_psnr_netG_A2B.pth',
                    )

                    if (
                        args.eval_with_ema
                        and ema_helpers
                    ):
                        apply_ema_to_generators(
                            ema_helpers,
                            models,
                        )

                        save_A_to_B_weights(
                            best_psnr_path,
                            models,
                        )

                        restore_generators_from_ema(
                            ema_helpers,
                            models,
                        )
                    else:
                        save_A_to_B_weights(
                            best_psnr_path,
                            models,
                        )

                if not improved:
                    best[
                        'bad_epochs'
                    ] += 1

                row = {
                    'epoch': epoch,
                    'lr_g': lr_g,
                    'lr_d': lr_d,
                    **avg,
                    'ssim_mean': cur_ssim,
                    'ssim_std': (
                        summary[
                            'case_ssim_std'
                        ]
                    ),
                    'psnr_mean': cur_psnr,
                    'psnr_std': (
                        summary[
                            'case_psnr_std'
                        ]
                    ),
                    'best_ssim': (
                        best[
                            'best_ssim'
                        ]
                    ),
                    'best_psnr': (
                        best[
                            'best_psnr'
                        ]
                    ),
                    'best_ssim_epoch': (
                        best[
                            'best_ssim_epoch'
                        ]
                    ),
                    'best_psnr_epoch': (
                        best[
                            'best_psnr_epoch'
                        ]
                    ),
                    'bad_epochs': (
                        best[
                            'bad_epochs'
                        ]
                    ),
                    'time_min': (
                        time.time()
                        - epoch_start
                    )
                    / 60.0,
                }

                append_csv(
                    metrics_csv,
                    row,
                )

                save_full_checkpoint(
                    os.path.join(
                        exp_dir,
                        'latest.pth',
                    ),
                    epoch,
                    global_step,
                    models,
                    optimizers,
                    schedulers,
                    scaler,
                    args,
                    best,
                    ema_helpers,
                )

                if (
                    epoch
                    % args.save_content_every
                    == 0
                ):
                    save_full_checkpoint(
                        os.path.join(
                            exp_dir,
                            (
                                f'content_epoch_'
                                f'{epoch:03d}.pth'
                            ),
                        ),
                        epoch,
                        global_step,
                        models,
                        optimizers,
                        schedulers,
                        scaler,
                        args,
                        best,
                        ema_helpers,
                    )

                if args.save_epoch_weights:
                    save_A_to_B_weights(
                        os.path.join(
                            exp_dir,
                            (
                                f'netG_A2B_epoch_'
                                f'{epoch:03d}.pth'
                            ),
                        ),
                        models,
                    )

                print(
                    f'[Epoch '
                    f'{epoch:03d}/'
                    f'{args.num_epoch}] '
                    f'lr_g={lr_g:.2e} '
                    f'lr_d={lr_d:.2e} '
                    f'G={avg["G_sum"]:.4f} '
                    f'Gcyc={avg["G_cycle"]:.4f} '
                    f'GL1={avg["G_L1"]:.4f} '
                    f'Gadv={avg["G_adv"]:.4f} '
                    f'GcycAdv={avg["G_cycle_adv"]:.4f} '
                    f'D={avg["D_diff"]:.4f} '
                    f'Dcyc={avg["D_cycle"]:.4f} '
                    f'R1={avg["R1"]:.4f} '
                    f'SSIM={cur_ssim:.5f}'
                    f'±{summary["case_ssim_std"]:.5f} '
                    f'PSNR={cur_psnr:.3f}'
                    f'±{summary["case_psnr_std"]:.3f} '
                    f'best_ssim='
                    f'{best["best_ssim"]:.5f}'
                    f'@{best["best_ssim_epoch"]} '
                    f'best_psnr='
                    f'{best["best_psnr"]:.3f}'
                    f'@{best["best_psnr_epoch"]} '
                    f'patience='
                    f'{best["bad_epochs"]}/'
                    f'{args.patience} '
                    f'time='
                    f'{(time.time() - epoch_start) / 60.0:.1f}m'
                )

            if is_main_process():
                stop_now = (
                    best[
                        'bad_epochs'
                    ]
                    >= args.patience
                )
            else:
                stop_now = False

            if (
                dist.is_available()
                and dist.is_initialized()
            ):
                stop_flag = torch.tensor(
                    [
                        1
                        if stop_now
                        else 0
                    ],
                    device=device,
                    dtype=torch.int64,
                )

                dist.broadcast(
                    stop_flag,
                    src=0,
                )

                stop_now = bool(
                    stop_flag.item()
                )

            if stop_now:
                if is_main_process():
                    print(
                        f'[EarlyStop] '
                        f'patience reached: '
                        f'{best["bad_epochs"]}/'
                        f'{args.patience}'
                    )

                break

    finally:
        cleanup_ddp()


if __name__ == '__main__':
    main()